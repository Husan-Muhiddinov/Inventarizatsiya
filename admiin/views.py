from django.shortcuts import redirect, render, get_object_or_404, HttpResponse
from django.http import HttpResponseBadRequest
from django.urls import reverse_lazy
from index.models import Category, Model, Product, Responsible, RoomsModel, User
from django.views.generic import ListView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import *
from django.db.models import Count, Q
from django.contrib.auth.decorators import login_required
from .freshed_list import remove_from_list_item
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook

def test(request):
    return render(request, 'building.html')

def page_not_found_view(request, exception):
    return render(request, 'pages-404.html', status=404)

def custom_error_view(request, exception=None):
    return render(request, "pages-500.html", status=500)

# def custom_permission_denied_view(request, exception=None):
#     return render(request, "403.html", {})

# def custom_bad_request_view(request, exception=None):
#     return render(request, "errors/400.html", {})

@login_required
def Admin_index(request):
    products = Product.objects.filter(group=request.user.groups.first()).filter(category_id__isnull=False)
    prod_for_count = Product.objects.filter(group=request.user.groups.first())
    product_count = len(prod_for_count)
    category_filt = Category.objects.filter(group=request.user.groups.first())
    category_count = category_filt.count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()
    categories = category_filt.annotate(count=Count('category'))
    query_count = products.filter(status=1).count()
    query_count_tw = products.filter(status=3).count()
    query_count_ir = products.filter(status=2).count()
    query_count_in = products.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(categories, 20)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'product_count': product_count,
        'query': query,
        'query_count': query_count, 
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'catalog_category.html', context=context)


class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CategoryCreateForm
    template_name = 'file/category-add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-categories')

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        context['category_count'] = Category.objects.filter(group=group).count()
        context['model_count'] = Model.objects.filter(group=group).count()
        context['res_count'] = Responsible.objects.filter(group=group).count()
        return context


class ResponsibleCreateView(LoginRequiredMixin, CreateView):
    model = Responsible
    form_class = ResponsibleCreateForm
    template_name = 'file/responsible_add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-responsibles')
    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        context['category_count'] = Category.objects.filter(group=group).count()
        context['model_count'] = Model.objects.filter(group=group).count()
        context['res_count'] = Responsible.objects.filter(group=group).count()
        return context


# class RoomsCreateView(LoginRequiredMixin, CreateView):
#     model = RoomsModel
#     form_class = RoomsForm
#     template_name = 'file/rooms_add.html'
#     login_url = 'login'
#     success_url = reverse_lazy('pages:rooms')

#     def form_valid(self, form):
#         form.instance.author = self.request.user
#         return super().form_valid(form)


class ModelCreateView(LoginRequiredMixin, CreateView):
    model = Model
    form_class = ModelCreateForm
    template_name = 'file/model_add.html'
    login_url = 'login'
    success_url = reverse_lazy('admin-models')

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.group = self.request.user.groups.first()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        group = self.request.user.groups.first()
        context['model_count'] = Model.objects.filter(group=group).count()
        context['category_count'] = Category.objects.filter(group=group).count()
        context['res_count'] = Responsible.objects.filter(group=group).count()
        return context
 

@login_required
def baseview(request, pk):
    q = Product.objects.filter(group=request.user.groups.first())
    prod_count = len(q)
    query_ = q.filter(category_id__id=pk).order_by('pk')
    get_cat = Category.objects.get(id=pk)
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'prod_count': prod_count,
        'query': query, 
        'get_cat': get_cat, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'index.html', context=context)

# PRODUCT BASED views

@login_required
def Unknown_Devices(request):
    q = Product.objects.filter(group=request.user.groups.first())
    query_ = q.filter(category_id=None)
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'query': query, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'unknown_devices.html', context=context)

@login_required
def EquipmentCreateView(request, pk):
    products_count = Product.objects.filter(group=request.user.groups.first()).count()
    category = Category.objects.get(id=pk)
    inputs = request.POST.getlist('inputs')
    request.POST._mutable = True
    request.POST['description'] = ' | '.join(inputs)
    request.POST._mutable = False
    form = EquipmentCreateForm(request.user)
    if request.method == 'POST':
        form = EquipmentCreateForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            equipment = Product(
                group=request.user.groups.first(),
                name=form.cleaned_data['name'],
                schet=form.cleaned_data['schet'],
                category_id=category,
                room_number=form.cleaned_data['room_number'],
                inventar_number=form.cleaned_data['inventar_number'],
                model_id=form.cleaned_data['model_id'],
                responsible_id=form.cleaned_data['responsible_id'],
                images=form.cleaned_data['image'],
                seria_number=form.cleaned_data['seria_number'],
                year_of_manufacture=form.cleaned_data['year_of_manufacture'],
                unit_of_measurement=form.cleaned_data['unit_of_measurement'],
                description=form.cleaned_data['description'],
            )
            equipment.save()  
            red = category.id
            return redirect('base-view', pk=red)
        # else:
        #     return ("""<h1 style="color: "cyan"; text-align: "center";">Forma validatsiyadan o'tmadi!!!</h1>""")
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'form': form, 
        'cat_id': category.id, 
        'products_count': products_count,
    }
    return render(request, 'file/product_add.html', context=context)


@login_required
def ProductDetailView(request, pk):
    eq = get_object_or_404(Product, pk=pk)
    descrip = eq.description.split(' | ')
    descript = remove_from_list_item(descrip, '|')
    form = ProductDetailUpdateForm(request.POST or None, instance=eq)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('product-detail', pk=eq.pk)
    
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'query': eq, 
        'descript': descript, 
        'form': form,
    }
    return render(request, 'file/product-detail.html', context=context)


@login_required
def ProductUpdateView(request, pk):
    equipment = get_object_or_404(Product, pk=pk)
    if equipment.category_id is not None:
        red = equipment.category_id.id
    form = ProductUpdateForm(request.user, instance=equipment)
    if request.method == 'POST':
        if 'inputs' and 'status' in request.POST:
            inputs = request.POST.getlist('inputs')
            request.POST._mutable = True
            request.POST['status'] = request.POST['status']
            request.POST['description'] = equipment.description + ' | ' + ' | '.join(inputs)
            request.POST['responsible_person'] = request.POST['responsible_person']
            request.POST._mutable = False
        form = ProductUpdateForm(request.user, request.POST or None, request.FILES or None, instance=equipment)
        if form.is_valid():
            form.save()
        else:
            return HttpResponse(form.errors)
        if equipment.category_id is not None:
            red = equipment.category_id.id
            return redirect('base-view', pk=red)
        else:
            return redirect('unknown-devices')
    
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()
    
    context = {
        'form': form, 
        'equ': equipment,
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
    }

    if equipment.category_id is not None:
        context['red'] = red
        return render(request, 'file/product_update.html', context=context)
    else:
        return render(request, 'file/product_update.html', context=context)


@login_required
def ProductDeleteView(request, pk):
    query = Product.objects.get(pk=pk)
    if query.category_id is not None:
        red = query.category_id.id
    if request:
        query.delete()
        if query.category_id is not None:
            return redirect('base-view', pk=red)
        else:
            return redirect('unknown-devices')
    return render(request, 'admin/admin-base.html')


# SEARCh views

class SearchResultsView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'search-results.html'
    # paginate_by = 2

    def get_queryset(self):
        if self.request.GET.get('search') is not None:
            query = self.request.GET.get('search')
        else:
            query = ''
        object_list = Product.objects.filter(group=self.request.user.groups.first()).filter(
            Q(inventar_number__icontains=query) | Q(name__icontains=query)
        ).order_by('pk')
        return object_list
    
    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        query_count = self.object_list.filter(status=1).count()
        query_count_tw = self.object_list.filter(status=3).count()
        query_count_ir = self.object_list.filter(status=2).count()
        query_count_in = self.object_list.filter(status=0).count()
        category_count = Category.objects.filter(group=self.request.user.groups.first()).count()
        model_count = Model.objects.filter(group=self.request.user.groups.first()).count()
        res_count = Responsible.objects.filter(group=self.request.user.groups.first()).count()
        context['query_count'] = query_count
        context['query_count_tw'] = query_count_tw
        context['query_count_ir'] = query_count_ir
        context['query_count_in'] = query_count_in
        context['category_count'] = category_count
        context['model_count'] = model_count
        context['res_count'] = res_count
        return context

# CATEGORIES views
    
@login_required
def AdminCategories(request):
    query_ = Category.objects.filter(group=request.user.groups.first()).order_by('pk')
    categ_count = len(query_)
    query_p = Product.objects.filter(group=request.user.groups.first())
    query_count = query_p.filter(status=1).count()
    query_count_tw = query_p.filter(status=3).count()
    query_count_ir = query_p.filter(status=2).count()
    query_count_in = query_p.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 20)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'categ_count': categ_count, 
        'query': query, 
        'query_count': query_count, 
        'query_count_tw': query_count_tw, 
        'query_count_ir': query_count_ir, 
        'query_count_in': query_count_in,
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
    }

    return render(request, 'card.html', context=context)


@login_required
def AdminCategoryEdit(request, pk):
    query = get_object_or_404(Category, pk=pk)
    form = CategoryEditForm(request.POST or None, request.FILES or None, instance=query)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('admin-categories')

    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'form': form,
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
    }
    return render(request, 'file/category_edit.html', context=context)


@login_required
def AdminCategoryDelete(request, pk):
    query = Category.objects.get(pk=pk)
    if request:
        query.delete()
        return redirect('admin-categories')
    return render(request, 'card.html')


# UNIVERSAL EXCEL UPLODER VIEW

@login_required
def UploadExcelFile(request):
    if request.method == "POST":
        user_gp = request.user.groups.first()
        form = UploadExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            if form.cleaned_data['excel_file_cat'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_cat'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    category_count = Category.objects.filter(group=user_gp)
                    if user_gp.max_limit_category == len(category_count):
                        error = _("Siz boshqa kategoriya qo'sha olmaysiz!")
                        return redirect("category-create", {'error': error})
                        break
                    else:
                        result = Category(
                            group = user_gp,
                            name_uz = i[0].value,
                            name_en = i[1].value,
                            name_ru = i[2].value,
                        )
                        result.save()
                        return redirect("admin-categories")
            elif form.cleaned_data['excel_file_mod'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_mod'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    model_count = Model.objects.filter(group=user_gp)
                    if user_gp.max_limit_model > len(model_count):
                        result = Model(
                            group = user_gp,
                            name = i[0].value,
                            description = i[1].value,
                        )
                        result.save()
                return redirect("admin-models")
            elif form.cleaned_data['excel_file_per'] is not None:
                wb = load_workbook(filename=request.FILES['excel_file_per'].file)
                worksheet = wb["Sheet1"]
                for i in worksheet.iter_rows(min_row=2):
                    res_count = Responsible.objects.filter(group=user_gp)
                    if user_gp.max_limit_responsible > len(res_count):
                        result = Responsible(
                            group = request.user.groups.first(),
                            fullname_uz = i[0].value,
                            fullname_en = i[1].value,
                            fullname_ru = i[2].value,
                            description_uz = i[3].value,
                            description_en = i[4].value,
                            description_ru = i[5].value,
                        )
                        result.save()
                return redirect("admin-responsibles")
            elif form.cleaned_data['excel_file_pro'] is not None:
                try:
                    wb = load_workbook(filename=request.FILES['excel_file_pro'].file)
                    worksheet = wb["Sheet1"]
                    for i in worksheet.iter_rows(min_row=2):
                        product_count = Product.objects.filter(group=user_gp)
                        if user_gp.max_limit_product > len(product_count):
                            result = Product(
                                group = request.user.groups.first(),
                                name = i[0].value,
                                schet = i[1].value,
                                inventar_number = i[2].value,
                                seria_number = i[3].value,
                                description = i[4].value,
                                year_of_manufacture = i[5].value,
                                unit_of_measurement = i[6].value,
                                room_number = i[7].value,
                            )
                            result.save()
                    return redirect("unknown-devices")
                except Exception as e:
                    error = _("Faylda bazada mavjud inventar raqamga ega jihoz mavjud!")
                    return render(request, 'file/product_add_excel.html', {'error': error, 'e': e})
            else:
                return HttpResponseBadRequest()
        else:
            return HttpResponseBadRequest()
    else:
        return render(request, 'unknown_devices.html', {})

def product_create_excel(request):
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
    }
    return render(request, 'file/product_add_excel.html', context=context)
    
# MODELS views

@login_required
def AdminModels(request):
    query_ = Model.objects.filter(group=request.user.groups.first()).order_by('pk')
    model_count = len(query_)
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()
    query_p = Product.objects.filter(group=request.user.groups.first())
    query_count = query_p.filter(status=1).count()
    query_count_tw = query_p.filter(status=3).count()
    query_count_ir = query_p.filter(status=2).count()
    query_count_in = query_p.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 20)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'model_count': model_count, 
        'query': query, 
        'query_count': query_count, 
        'query_count_tw': query_count_tw, 
        'query_count_ir': query_count_ir, 
        'query_count_in': query_count_in,
        'category_count': category_count,
        'res_count': res_count,
    }

    return render(request, 'models.html', context=context)


@login_required
def AdminModelEdit(request, pk):
    query = get_object_or_404(Model, pk=pk)
    form = ModelEditForm(request.POST or None, request.FILES or None, instance=query)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('admin-models')
    
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    context = {
        'category_count': category_count,
        'res_count': res_count,
        'model_count': model_count,
        'form': form,
    }
    return render(request, 'file/models_edit.html', context=context)


@login_required
def AdminModelDelete(request, pk):
    query = Model.objects.get(pk=pk)
    if request:
        query.delete()
        return redirect('pages:models')
    return render(request, 'models.html')

# RESPONSIBLE PERSONS views

@login_required
def AdminResponsibles(request):
    query_ = Responsible.objects.filter(group=request.user.groups.first()).order_by('pk')
    res_count = len(query_)
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    query_p = Product.objects.filter(group=request.user.groups.first())
    query_count = query_p.filter(status=1).count()
    query_count_tw = query_p.filter(status=3).count()
    query_count_ir = query_p.filter(status=2).count()
    query_count_in = query_p.filter(status=0).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'res_count': res_count, 
        'category_count': category_count, 
        'model_count': model_count, 
        'query': query, 
        'query_count': query_count, 
        'query_count_tw': query_count_tw, 
        'query_count_ir': query_count_ir, 
        'query_count_in': query_count_in,
    }
    return render(request, 'responsible_person.html', context=context)


@login_required
def AdminResponsibleEdit(request, pk):
    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()
    query = get_object_or_404(Responsible, pk=pk)
    form = ResponsibleEditForm(request.POST or None, instance=query)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('pages:responsible')

    context = {
        'form': form,
        'res_count': res_count, 
        'category_count': category_count, 
        'model_count': model_count,
    }
    return render(request, 'file/responsible-edit.html', context=context)


@login_required
def AdminResponsibleDelete(request, pk):
    query = Responsible.objects.get(pk=pk)
    if request:
        query.delete()
        return redirect('pages:responsible')
    return render(request, 'responsible_person.html')

def AdminResponsibleProducts(request, pk):
    q = Product.objects.filter(group=request.user.groups.first())
    prod_count = len(q)
    query_ = q.filter(responsible_id__id=pk).order_by('pk')
    get_cat = Responsible.objects.get(id=pk)
    query_count = query_.filter(status=1).count()
    query_count_tw = query_.filter(status=3).count()
    query_count_ir = query_.filter(status=2).count()
    query_count_in = query_.filter(status=0).count()

    category_count = Category.objects.filter(group=request.user.groups.first()).count()
    model_count = Model.objects.filter(group=request.user.groups.first()).count()
    res_count = Responsible.objects.filter(group=request.user.groups.first()).count()

    page = request.GET.get('page', 1)
    paginator = Paginator(query_, 50)
    try:
        query = paginator.page(page)
    except PageNotAnInteger:
        query = paginator.page(1)
    except EmptyPage:
        query = paginator.page(paginator.num_pages)

    context = {
        'category_count': category_count,
        'model_count': model_count,
        'res_count': res_count,
        'prod_count': prod_count,
        'query': query, 
        'get_cat': get_cat, 
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,
        }
    return render(request, 'index-responsible.html', context=context)

# ROOM views

@login_required
def AdminRooms(request):
    user = request.user.groups.first()
    category_count = Category.objects.filter(group=user).count()
    model_count = Model.objects.filter(group=user).count()
    res_count = Responsible.objects.filter(group=user).count()
    queryset = Product.objects.filter(group=user).order_by('room_number')
    room_list = []
    for u in queryset:
        if u.room_number not in room_list:
            room_list.append(u.room_number)
    context = {
        'user': user,
        'queryset': queryset,
        'room_list': room_list,
        'res_count': res_count, 
        'category_count': category_count, 
        'model_count': model_count,
    }
    return render(request, 'rooms_catalog.html', context=context)

@login_required
def AdminRoomFloor(request, pk):
    user = request.user.groups.first()
    category_count = Category.objects.filter(group=user).count()
    model_count = Model.objects.filter(group=user).count()
    res_count = Responsible.objects.filter(group=user).count()

    context = {
        'pka': pk,
        'res_count': res_count, 
        'category_count': category_count, 
        'model_count': model_count,
    }
    return render(request, 'rooms_catalog_floor.html', context=context)

@login_required
def AdminRoomDetail(request, slug):
    prod = request.user.groups.first()
    if prod.ui == True:
        query = Product.objects.filter(
            Q(room_number__icontains=slug)
        ).order_by('pk')
    else:
        query = Product.objects.filter(room_number=slug).order_by('pk')

    query_count = query.filter(status=1).count()
    query_count_tw = query.filter(status=3).count()
    query_count_ir = query.filter(status=2).count()
    query_count_in = query.filter(status=0).count()

    category_count = Category.objects.filter(group=prod).count()
    model_count = Model.objects.filter(group=prod).count()
    res_count = Responsible.objects.filter(group=prod).count()

    context = {
        'query': query,
        'query_count': query_count,
        'query_count_tw': query_count_tw,
        'query_count_ir': query_count_ir,
        'query_count_in': query_count_in,

        'res_count': res_count, 
        'category_count': category_count, 
        'model_count': model_count,
        }
    return render(request, 'file/room_devices.html', context=context)
